import os
from pathlib import Path
from openpyxl import Workbook
import zipfile

def create_test_xlsx():
    """Create a test XLSX file and analyze its structure"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Try different ways to set the cell value
    cell = ws['A1']
    cell.value = "Test"
    cell.data_type = 's'  # Explicitly set as string type

    wb.save("test_xlsx.xlsx")
    print("Created test_xlsx.xlsx")

    # Analyze the file structure
    with zipfile.ZipFile("test_xlsx.xlsx", 'r') as zf:
        print("ZIP file contents:")
        for name in zf.namelist():
            print(f"  {name}")
            if name.endswith('.xml'):
                try:
                    content = zf.read(name).decode('utf-8')
                    if 'sheet1' in name.lower():
                        print(f"\nContent of {name}:")
                        print(content[:500])  # First 500 chars
                        if 'sharedStrings' in content or 's="' in content:
                            print("WARNING: Found shared string references!")
                except Exception as e:
                    print(f"Error reading {name}: {e}")

    return "test_xlsx.xlsx"

def test_read_xlsx(file_path):
    """Test reading the XLSX file"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        print(f"Successfully loaded workbook: {ws.title}")

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    print(f"Cell {cell.coordinate}: {cell.value} (type: {cell.data_type})")
                else:
                    print(f"Cell {cell.coordinate}: None")

        wb.close()
        return True

    except Exception as e:
        print(f"Error reading XLSX: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    # Create and test XLSX file
    file_path = create_test_xlsx()
    success = test_read_xlsx(file_path)

    # Clean up
    if os.path.exists(file_path):
        os.remove(file_path)

    print(f"\nTest {'PASSED' if success else 'FAILED'}")